import os, random

from . import ledger
from ..employee import Employee, secure
from ..user.view import TokenIn, db
from database.sqlite_async import AsyncSqlite
import aiosqlite

from flask import render_template, url_for, redirect, flash
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image

from apiflask import Schema
from apiflask.fields import String
from apiflask.validators import OneOf

alignment = Alignment(horizontal="center")
border = Border(
    left=Side(border_style="thin", color="FF000000"),
    right=Side(border_style="thin", color="FF000000"),
    top=Side(border_style="thin", color="FF000000"),
    bottom=Side(border_style="thin", color="FF000000"),
)


async def get_which_workbook(raw_file_name, new_clue):
    file_name = None
    if "public" in raw_file_name:
        if os.path.exists(f"download/public_{new_clue}.xlsx"):
            file_name = f"download/public_{new_clue}.xlsx"
    if "personal" in raw_file_name:
        if os.path.exists(f"download/personal_{new_clue}.xlsx"):
            file_name = f"download/personal_{new_clue}.xlsx"
    workbook = load_workbook(filename=file_name or raw_file_name)
    return workbook, workbook.active


async def get_assert_basic(whole_info):
    assert_type = await db.select_db("category", "name", cid=whole_info["assert_type"])
    assert_type = assert_type[0][0]
    assert_admin = await db.select_db(
        "user", "username", user_id=int(whole_info["assert_admin"])
    )
    assert_admin = assert_admin[0][0]
    assert_id = whole_info["assert_type"] + str(random.random())[2:12]
    return assert_id, assert_type, assert_admin


class BaseAssert(Schema):
    assert_id = String(required=True)
    assert_name = String(required=True)
    assert_model = String(required=True)
    YoN = String(required=True)
    bought_date = String(required=True)
    assert_admin = String(required=True)
    submit = String(required=True)
    assert_type = String(required=True)


class PublicAssert(BaseAssert):
    # assert_type = String(required=True, validate=OneOf(
    #     ['房屋及建筑物', '办公家具', '电脑及打印机', '家电', '机械设备', '数码产品', '车辆', '其他']))
    TDM = String(required=True)


@ledger.get("/public")
@ledger.input(TokenIn, location="query")
async def public_show(query_data):
    curr_token = query_data["token"]
    curr_user = await Employee.get_user_by_token(curr_token)

    if curr_user.department_id:
        # curr_department_name = await Employee.get_department_by_id(int(curr_user.department_id))
        privileges = await curr_user.get_privileges(curr_token)
        if len(privileges) > 1:
            return redirect(
                url_for("admin.get_all_asserts_by_department", token=curr_token)
            )
        else:
            # workbook, sheet = await get_which_workbook('templates/public0.xlsx', curr_department_name)
            # table = []
            # for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3, max_col=sheet.max_column,
            #                            values_only=True):
            #     if None not in row:
            #         table.append(row)
            return render_template("public_assert.html", curr_user=curr_user)
    else:
        flash("请完善个人基本信息")
        return ""


@ledger.post("/public")
@ledger.input(PublicAssert, location="form")
@ledger.input(TokenIn, location="query")
async def public_post(data, query_data):
    curr_token = query_data["token"]
    curr_user = await Employee.get_user_by_token(curr_token)
    curr_department_name = await Employee.get_department_by_id(
        int(curr_user.department_id)
    )
    workbook, sheet = await get_which_workbook(
        "assert/templates/public0.xlsx", curr_department_name
    )
    rows = sheet.max_row
    rows += 1
    assert_id, assert_type, assert_admin = await get_assert_basic(data)
    msg = await db.upsert(
        "public_assert",
        {
            "aid": assert_id,
            "cid": data["assert_type"],
            "name": data["assert_name"],
            "model": data["assert_model"],
            "is_fixed": 1 if data["YoN"] == "True" else 0,
            "purchase_date": data["bought_date"],
            "manager": data["TDM"],
            "department_id": int(curr_user.department_id),
            "admin_id": data["assert_admin"],
        },
        0,
    )
    if not msg:
        flag = 1
    else:
        return msg
    # sql=f"insert into public_assert values {assert_id, int(data['assert_type']), data['assert_name'], data['assert_model'], 1 if data['YoN'] == 'True' else 0, data['bought_date'], data['TDM'], int(data['assert_admin'])}"
    # db.just_exe(sql)
    for l in range(len("CDEFGHIJ")):
        sheet["CDEFGHIJ"[l] + str(rows)] = [
            assert_type,
            assert_id,
            data["assert_name"],
            data["assert_model"],
            "是" if data["YoN"] == "True" else "否",
            data["bought_date"],
            assert_admin,
            data["TDM"],
        ][l]

        sheet["CDEFGHIJ"[l] + str(rows)].alignment = alignment
        sheet["CDEFGHIJ"[l] + str(rows)].border = border
    if curr_department_name not in sheet["D4"].value:
        sheet["D4"] = curr_department_name + sheet["D4"].value
    assert_admin = await db.select_db(
        "user", "username", user_id=int(data["assert_admin"])
    )
    sheet["H6"] = assert_admin[0][0]
    workbook.save(f"download/public_{curr_department_name}.xlsx")
    if flag:
        flash("已添加")
    return redirect(url_for("ledger.public_show", token=curr_token))


async def get_table_by_page(sheet, start_row, per, sum):
    if sum % per or sum / per:
        pages = int(sum / per) + 1 if per * int(sum / per) < sum else int(sum / per)
        start_page = [start_row + per * n for n in range(pages)]
        end_page = [n + per for n in start_page][:-1]
        end_page.append(start_row + sum)
        pagination = dict(zip(start_page, end_page))
        results = []
        for k, v in pagination.items():
            table = []
            for row in sheet.iter_rows(
                min_row=k,
                max_row=v,
                min_col=3,
                max_col=sheet.max_column,
                values_only=True,
            ):
                if None not in row:
                    table.append(row)
            results.append(table)
        return results, pages


@ledger.get("/personal/<int:page>")
@ledger.input(TokenIn, location="query")
async def personal_show(page, query_data):
    curr_token = query_data["token"]
    curr_user = await Employee.get_user_by_token(curr_token)
    workbook, sheet = await get_which_workbook(
        "assert/templates/personal0.xlsx", curr_user.username
    )

    table = []
    rows = await db.just_exe(
        f"select count(*) from personal_assert where personal_id={curr_user.user_id}"
    )
    rows = rows[0][0]
    pagination, pages = await get_table_by_page(sheet, 8, 9, rows)
    page_result = pagination[page - 1]
    import pprint

    pprint.pprint(page_result)
    return render_template(
        "personal_assert.html",
        table=page_result,
        curr_user=curr_user,
        pages=pages,
    )


@ledger.post("/personal")
@ledger.input(BaseAssert, location="form")
@ledger.input(TokenIn, location="query")
async def personal_post(data, query_data):
    curr_token = query_data["token"]
    curr_user = await Employee.get_user_by_token(curr_token)
    workbook, sheet = await get_which_workbook(
        "assert/templates/personal0.xlsx", curr_user.username
    )
    rows = sheet.max_row
    rows += 1
    assert_id, assert_type, assert_admin = await get_assert_basic(data)
    msg = await db.upsert(
        "personal_assert",
        {
            "aid": assert_id,
            "cid": data["assert_type"],
            "name": data["assert_name"],
            "model": data["assert_model"],
            "is_fixed": 1 if data["YoN"] == "True" else 0,
            "purchase_date": data["bought_date"],
            "admin_id": data["assert_admin"],
            "personal_id": curr_user.user_id,
        },
        0,
    )
    if not msg:
        flash("数据库已记录")
    else:
        return msg
    for l in range(len("CDEFGHI")):
        sheet["CDEFGHI"[l] + str(rows)] = [
            assert_type,
            assert_id,
            data["assert_name"],
            data["assert_model"],
            "是" if data["YoN"] == "True" else "否",
            data["bought_date"],
            assert_admin,
        ][l]

        sheet["CDEFGHI"[l] + str(rows)].alignment = alignment
        sheet["CDEFGHI"[l] + str(rows)].border = border
    sheet["D4"] = sheet["D4"].value.replace("个人", curr_user.username)
    assert_admin = await db.select_db(
        "user", "username", user_id=int(data["assert_admin"])
    )
    sheet["H6"] = assert_admin[0][0]
    workbook.save(f"download/personal_{curr_user.username}.xlsx")
    return redirect(url_for("ledger.personal_show", token=curr_token))
