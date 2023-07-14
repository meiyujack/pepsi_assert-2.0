import os
import re
from markupsafe import Markup

from . import admin
from ..employee import db, Employee
from ..user.view import TokenIn
from ..ledger.view import get_which_workbook, alignment
from auth import WebSecurity

from apiflask import Schema
from apiflask.fields import String
from flask import send_file, json, render_template, flash, url_for, redirect, request
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

secure = WebSecurity('ocefjVp_pL4Iens21FTjsA')


class DownloadIn(TokenIn):
    department = String(required=True)


class AlterIn(TokenIn):
    uid = String(required=True)
    rid = String(required=True)


class DeletePersonalAssert(TokenIn):
    aid = String(required=True)


class DeletePublicAssert(DeletePersonalAssert):
    department = String(required=True)


class UpdatePersonalAssert(DeletePersonalAssert):
    user = String(required=True)


class UpdatePublicAssert(DeletePersonalAssert):
    department=String(required=True)


class UpdatePersonalPost(Schema):
    assert_type = String(required=True)
    assert_name = String(required=True)
    assert_model = String(required=True)
    YoN = String(required=True)
    bought_date = String(required=True)
    assert_admin = String(required=True)


class UpdatePublicPost(UpdatePersonalPost):
    TDM=String(required=True)


def get_accurate_file(path, pattern):
    result = []
    for root, dirs, files in os.walk(path):
        for file in files:
            result.append(os.path.join(root, file))
    for i in result[:]:
        if not re.match(pattern, i):
            result.remove(i)
    return result


async def get_users_by_departments(departments_ids):
    result = []
    for i in departments_ids:
        users = await db.select_db('user', 'username', department_id=i)
        for user in users:
            if user[0]:
                result.append(user[0])
    return result


@admin.get('/get_privileges')
@admin.input(TokenIn, location='query')
async def get_privileges_by_token(data):
    token = data["token"]
    curr_user = await Employee.get_user_by_token(token)
    if curr_user.user_id == 104900 and curr_user.role_id == 8:
        await db.upsert("user",
                        {"user_id": curr_user.user_id, "username": curr_user.username, "password": curr_user.password,
                         "role_id": 8}, 0)

    privileges = await curr_user.get_privileges(token)
    result = []
    for privilege in privileges:
        result.append(privilege[0])
    return result


@admin.get('/')
@admin.input(TokenIn, location='query')
async def get_all_users(data):
    token = data["token"]
    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'find' in privilege:
            # users = await db.just_exe(
            #     'SELECT u.user_id,u.username,d.department_name,r.role_id,r.comment from "user" u join department d ,"role" r on u.department_id =d.department_id and u.role_id =r.role_id;')
            users = await db.just_exe(
                'SELECT u.user_id,u.username,u.department_id,r.role_id,r.comment from "user" u join "role" r on u.role_id=r.role_id')
            departments = []
            results = {}
            results["未选定部门"] = []
            # if users:
            for user in users:
                if user[2]:
                    department_name = await db.select_db("department", "department_name", department_id=user[2])
                    department_name = department_name[0][0]
                    if department_name not in departments:
                        departments.append(department_name)
                        results[department_name] = [
                            {"user_id": user[0], "username": user[1], "rname": user[4], "rid": user[3]}, ]
                    else:
                        results[department_name].append(
                            {"user_id": user[0], "username": user[1], "rname": user[4], "rid": user[3]}, )
                else:
                    results["未选定部门"].append(
                        {"user_id": user[0], "username": user[1], "rid": user[3], "rname": user[4]})
            # else:
            #     unallocated_users = await db.just_exe(
            #         'SELECT u.user_id,u.username,r.role_id,r.comment from "user" u join "role" r on u.role_id=r.role_id where u.department_id is NULL;')
            #     results["未选定部门"] = []
            #     # users = await db.select_db("user")
            #     for user in unallocated_users:
            #         results["未选定部门"].append({"user_id": user[0], "username": user[1],"rid":user[2],"rname":user[3]})

            return render_template("admin_employers.html", tables=results, curr_user=curr_user)
    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.get('/department_asserts')
@admin.input(TokenIn, location='query')
async def get_all_asserts_by_department(data):
    token = data["token"]
    rid = secure.get_info_by_token(token, 'rid')
    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    files = get_accurate_file("download/", 'download/public_*')
    if rid in (8,4):
        result = {}
        num = 0
        if len(files)==0:
            return '暂无人员添加公共资产信息'
        for file in files:
            workbook = load_workbook(file)
            sheet = workbook.active
            table = []
            for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                        max_col=sheet.max_column,
                                        values_only=True):
                if None not in row:
                    table.append(row)
            new_table = []
            new_table.append(table[0].__add__(('操作',)))
            department=re.match('^download/public_(.*).xlsx', file).group(1)
            result[department] = table
            if rid==8:
                for t in table[1:]:
                    new_table.append(
                        t.__add__((Markup(
                            f'<button type="button" name="update"><a href="update_public?aid={t[1]}&department={department}&token={token}">修改</a></button>'),
                                    Markup(
                                        f'<button type="button" name="remove"><a href="delete_public?aid={t[1]}&department={department}&token={token}")">删除</a></button>'))))
            else:
                for t in table[1:]:
                    new_table.append(
                        t.__add__((
                                    Markup(
                                        f'<button type="button" name="remove"><a href="delete_public?aid={t[1]}&department={department}&token={token}")">删除</a></button>'),)))
            result[department] = new_table
                
        return render_template('admin_public.html', tables=result, curr_user=curr_user)
    else:
        for privilege in privileges:
            if 'query' in privilege[0]:
                if curr_user.username == '汪鸿':
                    if os.path.exists('download/public_生产部.xlsx'):
                        workbook = load_workbook('download/public_生产部.xlsx')
                        sheet = workbook.active
                        result={}
                        table = []
                        for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                                   max_col=sheet.max_column,
                                                   values_only=True):
                            if None not in row:
                                table.append(row)
                        new_table = [table[0].__add__(('操作',))]
                        for t in table[1:]:
                            new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_public?aid={t[1]}&department=生产部&token={token}">修改</a></button>'),)))
                        result["生产部"]=new_table
                        return render_template('admin_public.html', tables=result, curr_user=curr_user)
                    return '暂无人员添加生产部公共资产信息'
                elif curr_user.username == '赵攀':
                    files = ['财务部', '人事部', '市场部', '物流部', '行政部']
                    result = {}
                    num = 0
                    for file in files:
                        if not os.path.exists(f'download/public_{file}.xlsx'):
                            num += 1
                            if num == len(files):
                                return '暂无人员添加相关部门公共资产信息'
                        else:
                            workbook = load_workbook(f'download/public_{file}.xlsx')
                            sheet = workbook.active
                            table = []
                            for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                                       max_col=sheet.max_column,
                                                       values_only=True):
                                if None not in row:
                                    table.append(row)
                            new_table = [table[0].__add__(('操作',))]
                            for t in table[1:]:
                                new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_public?aid={t[1]}&department={file}&token={token}">修改</a></button>'),)))
                            result[file] = new_table
                    return render_template('admin_public.html', tables=result, curr_user=curr_user)

                elif curr_user.username == '冯倩':
                    files = ['宜昌所', '恩施办', '襄阳所', '十堰办', '荆门办', '特渠部']
                    result = {}
                    num = 0
                    for file in files:
                        if not os.path.exists(f'download/public_{file}.xlsx'):
                            num += 1
                            if num == len(files):
                                return '暂无人员添加相关部门公共资产信息'
                        else:
                            workbook = load_workbook(f'download/public_{file}.xlsx')
                            sheet = workbook.active
                            table = []
                            for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                                       max_col=sheet.max_column,
                                                       values_only=True):
                                if None not in row:
                                    table.append(row)
                            new_table = [table[0].__add__(('操作',))]
                            for t in table[1:]:
                                new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_public?aid={t[1]}&department={file}&token={token}">修改</a></button>'),)))
                            result[file] = new_table
                            return render_template('admin_public.html', tables=result, curr_user=curr_user)
                    return '暂无人员添加相关部门公共资产信息'
                else:
                    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.get('/personal_asserts')
@admin.input(TokenIn, location='query')
async def get_all_asserts_by_personal(data):
    token = data["token"]
    curr_user = await Employee.get_user_by_token(token)
    rid = secure.get_info_by_token(token, 'rid')

    if rid == 1:
        return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)
    else:
        if curr_user.username == '赵攀':
            users = await get_users_by_departments([2, 3, 4, 5, 6])
            result = {}
            num = 0
            for user in users:
                if not os.path.exists(f'download/personal_{user}.xlsx'):
                    num += 1
                    if num == len(users):
                        return '暂无人员添加个人资产信息'
                else:
                    workbook = load_workbook(f'download/personal_{user}.xlsx')
                    sheet = workbook.active
                    table = []
                    for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                               max_col=sheet.max_column,
                                               values_only=True):
                        if None not in row:
                            table.append(row)
                    new_table = [table[0].__add__(('操作',))]
                    for t in table[1:]:
                        new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_personal?aid={t[1]}&user={user}&token={token}">修改</a></button>'),)))
                    result[user] = new_table
            return render_template('admin_personal.html', tables=result, curr_user=curr_user)
        if curr_user.username == '汪鸿':
            users = await get_users_by_departments([1])
            result = {}
            num = 0
            for user in users:
                if not os.path.exists(f'download/personal_{user}.xlsx'):
                    num += 1
                    if num == len(users):
                        return '暂无人员添加个人资产信息'
                else:
                    workbook = load_workbook(f'download/personal_{user}.xlsx')
                    sheet = workbook.active
                    table = []
                    for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                               max_col=sheet.max_column,
                                               values_only=True):
                        if None not in row:
                            table.append(row)
                    new_table = [table[0].__add__(('操作',))]
                    for t in table[1:]:
                        new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_personal?aid={t[1]}&user={user}&token={token}">修改</a></button>'),)))
                    result[user] = new_table
            return render_template('admin_personal.html', tables=result, curr_user=curr_user)
        if curr_user.username == '冯倩':
            users = await get_users_by_departments([7, 8, 9, 10, 11, 12])
            result = {}
            num = 0
            for user in users:
                if not os.path.exists(f'download/personal_{user}.xlsx'):
                    num += 1
                    if num == len(users):
                        return '暂无人员添加个人资产信息'
                else:
                    workbook = load_workbook(f'download/personal_{user}.xlsx')
                    sheet = workbook.active
                    table = []
                    for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                               max_col=sheet.max_column,
                                               values_only=True):
                        if None not in row:
                            table.append(row)
                    new_table = [table[0].__add__(('操作',))]
                    for t in table[1:]:
                        new_table.append(t.__add__((Markup(
        f'<button type="button" name="alter"><a href="update_personal?aid={t[1]}&user={user}&token={token}">修改</a></button>'),)))
                    result[user] = new_table
            return render_template('admin_personal.html', tables=result, curr_user=curr_user)
    if rid in (8,4):
        users = await db.select_db('user', 'username')
        result = {}
        num = 0
        for user in users:
            if not os.path.exists(f'download/personal_{user[0]}.xlsx'):
                num += 1
                if num == len(users):
                    return '暂无人员添加个人资产信息'
            else:
                workbook = load_workbook(f'download/personal_{user[0]}.xlsx')
                sheet = workbook.active
                table = []
                for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=3,
                                           max_col=sheet.max_column,
                                           values_only=True):
                    if None not in row:
                        table.append(row)
                new_table = []
                new_table.append(table[0].__add__(('操作',)))
                if rid==8:
                    for t in table[1:]:
                        new_table.append(
                            t.__add__((Markup(
                                f'<button type="button" name="update"><a href="update_personal?aid={t[1]}&user={user[0]}&token={token}">修改</a></button>'),
                                       Markup(
                                           f'<button type="button" name="remove"><a href="delete_personal?aid={t[1]}&token={token}")">删除</a></button>'))))
                else:
                    for t in table[1:]:
                        new_table.append(
                            t.__add__((
                                       Markup(
                                           f'<button type="button" name="remove"><a href="delete_personal?aid={t[1]}&token={token}")">删除</a></button>'),)))

                result[user[0]] = new_table
        return render_template('admin_personal.html', tables=result, curr_user=curr_user)


@admin.get('/download')
@admin.input(DownloadIn, location='query')
async def download(data):
    token = data["token"]
    department = data["department"]
    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'download' in privilege[0]:
            if os.path.exists(f'download/public_{department}.xlsx'):
                wb, sheet = await get_which_workbook(f'download/public_{department}.xlsx', department)
                for n in range(sheet.max_row - 9):
                    sheet[f'B{10 + n}'] = str(n + 1)
                    sheet[f'B{10 + n}'].alignment = alignment
                wb.save(f'download/public_{department}.xlsx')
                return send_file(path_or_file=f'../download/public_{department}.xlsx')
            if os.path.exists(f'download/personal_{department}.xlsx'):
                wb, sheet = await get_which_workbook(f'download/personal_{department}.xlsx', department)
                for n in range(sheet.max_row - 9):
                    sheet[f'B{10 + n}'] = str(n + 1)
                    sheet[f'B{10 + n}'].alignment = alignment
                wb.save(f'download/personal_{department}.xlsx')
                return send_file(path_or_file=f'../download/personal_{department}.xlsx')
    else:
        return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.get('/alter_privilege')
@admin.input(AlterIn, location='query')
async def alter_privilege(data):
    token = data["token"]
    wanna_uid = data["uid"]
    wanna_rid = data["rid"]
    rid = secure.get_info_by_token(token, 'rid')
    uname = await db.select_db("user", "username", user_id=wanna_uid)
    uname = uname[0][0]
    rname = await db.select_db("role", "comment", role_id=wanna_rid)
    rname = rname[0][0]
    if rid == 8:
        msg = await db.just_exe(f"update user set role_id={wanna_rid} where user_id={wanna_uid};")
        if not msg:
            flash(f"成功修改{uname}的角色权限为{rname}")
            return redirect(url_for('admin.get_all_users', token=token))
        return msg
    return flash(f"无权限")


@admin.get('/delete_personal')
@admin.input(DeletePersonalAssert, location='query')
async def delete_personal(data):
    token = data["token"]
    aid = data["aid"]

    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)

    uid = await db.select_db("personal_assert", "personal_id", aid=aid)
    if uid:
        uid = uid[0][0]
        uname = await db.select_db("user", "username", user_id=uid)
        uname = uname[0][0]
        await db.just_exe(f'delete from personal_assert where aid = {aid};')
        for privilege in privileges:
            if 'delete' in privilege[0]:
                if os.path.exists(f'download/personal_{uname}.xlsx'):
                    workbook = load_workbook(f'download/personal_{uname}.xlsx')
                    await delete_from_excel(workbook, aid, uname)
                    flash("该行资产删除成功～")
                    return redirect(url_for("admin.get_all_asserts_by_personal", token=token))
        return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.get('/delete_public')
@admin.input(DeletePublicAssert,location='query')
async def delete_public(data):
    token = data["token"]
    aid = data["aid"]
    department=data.get("department")

    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    
    if department:
        department_id=await db.select_db("department","department_id",department_name=department)
    if department_id:
        department_id=department_id[0][0]
        department_name=await db.select_db("department","department_name",department_id=department_id)
        department_name=department_name[0][0]
        await db.just_exe(f'delete from public_assert where aid={aid};')
        for privilege in privileges:
            if 'delete' in privilege[0]:
                if os.path.exists(f'download/public_{department_name}.xlsx'):
                    workbook = load_workbook(f'download/public_{department_name}.xlsx')
                    await delete_from_excel(workbook, aid, department_name)
                    flash("该行资产删除成功～")
                    return redirect(url_for("admin.get_all_asserts_by_department", token=token))
        return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


async def delete_from_excel(workbook, aid, uname):
    sheet = workbook.active
    for r in range(1, sheet.max_row + 1):
        if sheet['D' + str(r)].value == aid:
            Worksheet.delete_rows(sheet, r)
            break
    if uname[-1] not in ['部','所','办']: 
        workbook.save(f'download/personal_{uname}.xlsx')
    else:
        workbook.save(f'download/public_{uname}.xlsx')


async def update_from_excel(workbook, aid, uname, data):
    
        sheet = workbook.active
        assert_type = await db.select_db('category', 'name', cid=int(data['assert_type']))
        assert_type = assert_type[0][0]
        assert_admin = await db.select_db('user', 'username', user_id=int(data['assert_admin']))
        assert_admin = assert_admin[0][0]
        if uname[-1] not in ['部','所','办']:
            for r in range(1, sheet.max_row + 1):
                if sheet['D' + str(r)].value == aid:        
                    for v in range(len("CEFGHI")):
                        sheet["CEFGHI"[v] + str(r)] = \
                        [assert_type, data["assert_name"], data["assert_model"], data["YoN"], data["bought_date"],
                        assert_admin][v]
            workbook.save(f'download/personal_{uname}.xlsx')
        else:
            for r in range(1, sheet.max_row + 1):
                if sheet['D' + str(r)].value == aid:        
                    for v in range(len("CEFGHIJ")):
                        sheet["CEFGHIJ"[v] + str(r)] = \
                        [assert_type, data["assert_name"], data["assert_model"], data["YoN"], data["bought_date"],
                        assert_admin,data['TDM']][v]
            workbook.save(f'download/public_{uname}.xlsx')
        
    

@admin.get('/update_personal')
@admin.input(UpdatePersonalAssert, location='query')
async def update_personal(data):
    token = data["token"]
    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'update' in privilege[0]:
            user = data["user"]
            aid = data["aid"]
            info = await db.select_db("personal_assert", aid=aid)
            if info:
                info = info[0]
            return render_template("admin_update_personal_asserts.html", curr_user=curr_user, user=user, aid=aid, info=info,next="personal_asserts?"+request.full_path.split('&')[-1])
    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.post('/update_personal')
@admin.input(UpdatePersonalPost, location='form')
@admin.input(UpdatePersonalAssert, location='query')
async def update_personal_assert(data, query_data):
    aid = query_data["aid"]
    info = await db.select_db("personal_assert", aid=aid)
    if info:
        info = info[0]
    token = query_data["token"]
    user = query_data["user"]

    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'update' in privilege[0]:

            update_info = {}
            if int(data["assert_type"]) != info[1]:
                update_info["cid"] = int(data["assert_type"])
            if data["assert_name"] != info[2]:
                update_info["name"] = data["assert_name"]
            if data['assert_model'] != info[3]:
                update_info["model"] = data["assert_model"]
            if data["YoN"] != info[4]:
                update_info["is_fixed"] = data["YoN"]
            if data["bought_date"] != info[5]:
                update_info["purchase_date"] = data["bought_date"]
            if data["assert_admin"] != info[6]:
                update_info["admin_id"] = data["assert_admin"]

            s = ""
            if update_info:
                for k, v in update_info.items():
                    s += f"{k}='{v}',"
                s = s[:-1]
                msg = await db.just_exe(f"update personal_assert set {s} where aid = {aid}")
                if msg:
                    return msg

                if os.path.exists(f'download/personal_{user}.xlsx'):
                    workbook = load_workbook(f'download/personal_{user}.xlsx')
                    await update_from_excel(workbook, aid, user, data)

                    flash("修改成功～")
                    return redirect(url_for("admin.get_all_asserts_by_personal", token=token))
            flash("当前没有修改～")
            return redirect(url_for("admin.update",aid=aid,user=user,token=token))

    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.get('/update_public')
@admin.input(UpdatePublicAssert, location='query')
async def update_public(data):
    token = data["token"]
    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'update' in privilege[0]:
            department = data["department"]
            aid = data["aid"]
            info = await db.select_db("public_assert", aid=aid)
            if info:
                info = info[0]
            return render_template("admin_update_public_asserts.html", curr_user=curr_user, department=department, aid=aid, info=info,next="department_asserts?"+request.full_path.split('&')[-1])
    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)


@admin.post('/update_public')
@admin.input(UpdatePublicPost,location='form')
@admin.input(UpdatePublicAssert,location='query')
async def update_public_assert(data,query_data):
    aid = query_data["aid"]
    info = await db.select_db("public_assert", aid=aid)
    if info:
        info = info[0]
    token = query_data["token"]
    department = query_data["department"]

    curr_user = await Employee.get_user_by_token(token)
    privileges = await curr_user.get_privileges(token)
    for privilege in privileges:
        if 'update' in privilege[0]:

            update_info = {}
            if int(data["assert_type"]) != info[1]:
                update_info["cid"] = int(data["assert_type"])
            if data["assert_name"] != info[2]:
                update_info["name"] = data["assert_name"]
            if data['assert_model'] != info[3]:
                update_info["model"] = data["assert_model"]
            if data["YoN"] != info[4]:
                update_info["is_fixed"] = data["YoN"]
            if data["bought_date"] != info[5]:
                update_info["purchase_date"] = data["bought_date"]
            if data["assert_admin"] != info[7]:
                update_info["admin_id"] = data["assert_admin"]
            if data["TDM"]!=info[6]:
                update_info["TDM"]=data["TDM"]
            s = ""
            if update_info:
                for k, v in update_info.items():
                    s += f"{k}='{v}',"
                s = s[:-1]
                msg = await db.just_exe(f"update public_assert set {s} where aid = {aid}")
                if msg:
                    return msg

                if os.path.exists(f'download/public_{department}.xlsx'):
                    workbook = load_workbook(f'download/public_{department}.xlsx')
                    await update_from_excel(workbook, aid, department, data)

                    flash("修改成功～")
                    return redirect(url_for("admin.get_all_asserts_by_department", token=token))
            flash("当前没有修改～")
            return redirect(url_for("admin.update",aid=aid,user=department,token=token))

    return json.dumps({"402": "你没有权限！"}, ensure_ascii=False)



